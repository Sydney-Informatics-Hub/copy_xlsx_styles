"""This copies the styling from one sheet to another.
"""

# TODO: copy named styles
# TODO: copy data validation
# TODO: Handle alternating styles
# TODO: copy auxiliary sheets

import argparse
from copy import copy

from openpyxl import load_workbook
from openpyxl.worksheet.dimensions import RowDimension, ColumnDimension
from openpyxl.utils.cell import get_column_letter

UNSUPPORTED_ATTRS = {'style', 'customFormat', 's', 'customHeight', 'customWidth'}
CELL_ATTRS = list({'font', 'fill', 'border', 'number_format', 'protection', 'alignment', 'style'} - UNSUPPORTED_ATTRS)
ROW_ATTRS = CELL_ATTRS + list(set(RowDimension.__fields__) - UNSUPPORTED_ATTRS)
COL_ATTRS = CELL_ATTRS + list(set(ColumnDimension.__fields__) - UNSUPPORTED_ATTRS)


def get_worksheet_for_path(sheet_path):
    if "!" not in sheet_path:
        return load_workbook(sheet_path).worksheets[0]
    path, _, sheet = sheet_path.rpartition("!")
    return load_workbook(path)[sheet]


def copy_styles(style_sheet, data_sheet):
    max_matched_row = min(style_sheet.max_row, data_sheet.max_row)
    max_matched_col = min(style_sheet.max_column, data_sheet.max_column)

    def generate_cell_pairs():
        for style_row, data_row in zip(style_sheet.iter_rows(max_row=max_matched_row, max_col=max_matched_col),
                                       data_sheet.iter_rows(max_row=max_matched_row, max_col=max_matched_col)):
            yield from zip(style_row, data_row)

        if data_sheet.max_column > style_sheet.max_column:
            for style_row, data_row in zip(style_sheet.iter_rows(max_row=max_matched_row, min_col=max_matched_col, max_col=max_matched_col),
                                           data_sheet.iter_rows(max_row=max_matched_row, min_col=max_matched_col + 1)):
                style_cell, = style_row
                for data_cell in data_row:
                    yield style_cell, data_cell

        if data_sheet.max_row > style_sheet.max_row:
            style_row, = style_sheet.iter_rows(min_row=max_matched_row, max_col=max_matched_col)
            for data_row in data_sheet.iter_rows(min_row=max_matched_row + 1, max_col=max_matched_col):
                yield from zip(style_row, data_row)

        if data_sheet.max_row > style_sheet.max_row and data_sheet.max_column > style_sheet.max_column:
            style_row, = style_sheet.iter_rows(min_row=max_matched_row, max_col=max_matched_col)
            style_cell, = style_row
            for data_row in data_sheet.iter_rows(min_row=max_matched_row + 1, min_col=max_matched_col + 1):
                for data_cell in data_row:
                    yield style_cell, data_cell

    for style_cell, data_cell in generate_cell_pairs():
        if style_cell.has_style:
            for attr in CELL_ATTRS:
                setattr(data_cell, attr, copy(getattr(style_cell, attr)))

    data_sheet.freeze_panes = style_sheet.freeze_panes

    for i in range(1, data_sheet.max_row):
        data_dim = data_sheet.row_dimensions[i]
        style_dim = style_sheet.row_dimensions[min(i, style_sheet.max_row)]
        for attr in ROW_ATTRS:
            setattr(data_dim, attr, copy(getattr(style_dim, attr)))

    for i in range(1, data_sheet.max_column):
        data_dim = data_sheet.column_dimensions[get_column_letter(i)]
        style_dim = style_sheet.column_dimensions[get_column_letter(min(i, style_sheet.max_column))]
        for attr in COL_ATTRS:
            setattr(data_dim, attr, copy(getattr(style_dim, attr)))

    data_sheet.auto_filter = style_sheet.auto_filter
    data_sheet.conditional_formatting = style_sheet.conditional_formatting


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("style_worksheet", type=get_worksheet_for_path)
    ap.add_argument("data_worksheet", type=get_worksheet_for_path)
    ap.add_argument("output_xlsx")
    args = ap.parse_args()
    copy_styles(args.style_worksheet, args.data_worksheet)
    args.data_worksheet.parent.save(args.output_xlsx)


if __name__ == '__main__':
    main()
